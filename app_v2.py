"""
🚀 Sistema de Finanças - Versão Simples
"""
from flask import Flask, render_template, request, jsonify, make_response
from flask_cors import CORS
import json
from datetime import datetime
from dateutil.relativedelta import relativedelta
import uuid
import os
import urllib.parse

app = Flask(__name__)
CORS(app)

# Arquivo para armazenar dados
DADOS_FILE = '/tmp/financas_dados.json'

def carregar_dados():
    """Carrega dados do arquivo"""
    if os.path.exists(DADOS_FILE):
        try:
            with open(DADOS_FILE, 'r') as f:
                dados = json.load(f)
                
                # Migrar categorias antigas (strings) para nova estrutura (dicts)
                if 'categorias' in dados and len(dados['categorias']) > 0:
                    if isinstance(dados['categorias'][0], str):
                        # Converte categorias antigas para nova estrutura
                        categorias_antigas = dados['categorias']
                        dados['categorias'] = [
                            {
                                'id': i + 1,
                                'nome': cat,
                                'tipo': 'DESPESA',  # Padrão para antigas
                                'subcategorias': []
                            }
                            for i, cat in enumerate(categorias_antigas)
                        ]
                        salvar_dados(dados)
                
                # Inicializa cartoes se não exists
                if 'cartoes' not in dados:
                    dados['cartoes'] = []
                
                # Inicializa faturas se não exists
                if 'faturas' not in dados:
                    dados['faturas'] = []
                
                # Se não tem categorias, criar padrão
                if 'categorias' not in dados or len(dados['categorias']) == 0:
                    dados['categorias'] = [
                        {
                            'id': 1,
                            'nome': 'Animal_Estimação',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Creche', 'Pet Shop', 'Ração', 'Remédio', 'Veterinario']
                        },
                        {
                            'id': 2,
                            'nome': 'Alimentação',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Almoco Trabalho', 'Doces', 'Feira', 'Lanche', 'paes', 'Pastel', 'Pizza', 'Restaurante', 'Supermercado']
                        },
                        {
                            'id': 3,
                            'nome': 'Casa',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Brastemp', 'Celular', 'Cloud', 'Comgas', 'Complemento Mensal', 'condominio', 'Eletro Eletronicos', 'Energia Eletrica', 'Faxineira', 'Internet', 'IPTU', 'Manutencao', 'Seguro']
                        },
                        {
                            'id': 4,
                            'nome': 'Cuidados_Pessoais',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Barba', 'Cabelo', 'Cosmeticos', 'Perfume']
                        },
                        {
                            'id': 5,
                            'nome': 'Educação',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Curso', 'Formatura', 'Lanche', 'material escolar', 'Mensalidade', 'Transporte', 'Vestibular']
                        },
                        {
                            'id': 6,
                            'nome': 'Eletronicos',
                            'tipo': 'DESPESA',
                            'subcategorias': ['App', 'Audio / Video', 'pilhas']
                        },
                        {
                            'id': 7,
                            'nome': 'Emprestimos',
                            'tipo': 'DESPESA',
                            'subcategorias': []
                        },
                        {
                            'id': 8,
                            'nome': 'Taxas_Impostos',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Anuidade', 'IOF']
                        },
                        {
                            'id': 9,
                            'nome': 'Investimento',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Açoes', 'Capitalização', 'CDB', 'Previdencia Privada', 'Renda Fixa', 'Tesouro Direto', 'Dólar']
                        },
                        {
                            'id': 10,
                            'nome': 'Jogos',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Aposta', 'Loteria', 'Rifa']
                        },
                        {
                            'id': 11,
                            'nome': 'Lazer',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Academia', 'Balada', 'Bebida', 'Churrasco', 'Cinema', 'Doces', 'Festa', 'Formatura', 'Futebol', 'Show', 'Streaming', 'Video game']
                        },
                        {
                            'id': 12,
                            'nome': 'Mesada',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Gustavo', 'Leonardo']
                        },
                        {
                            'id': 13,
                            'nome': 'Outros',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Não sei', 'Terceiros']
                        },
                        {
                            'id': 14,
                            'nome': 'Presentes',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Aniversario', 'Claudia', 'Dia das maes', 'Dia dos pais', 'Eduardo', 'Gustavo', 'Leonardo', 'Natal', 'Outros']
                        },
                        {
                            'id': 15,
                            'nome': 'Saude',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Convenio medico', 'dentista', 'estacionamento', 'farmacia', 'Oculos', 'Suplementacao']
                        },
                        {
                            'id': 16,
                            'nome': 'Transporte',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Acessorios Carro', 'Aquisicao Veiculo', 'Combustivel', 'estacionamento', 'garagem', 'IPVA / Licenciamento', 'Lavagem', 'manutencao', 'multas', 'pedagio', 'Seguro', 'Transp Publico', 'Uber']
                        },
                        {
                            'id': 17,
                            'nome': 'Vestuario',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Claudia', 'Eduardo', 'Familia', 'Gustavo', 'Leonardo']
                        },
                        {
                            'id': 18,
                            'nome': 'Viagem',
                            'tipo': 'DESPESA',
                            'subcategorias': ['Alimentacao', 'Animal Estimacao', 'Carro', 'Combustivel', 'Dolar', 'Farmacia', 'Hospedagem', 'Passagem', 'Passeio', 'pedagios', 'pontos / milhas', 'Passaporte', 'Presentes', 'Seguro', 'salaVip', 'Visto']
                        },
                        {
                            'id': 19,
                            'nome': 'Depositos',
                            'tipo': 'RECEITA',
                            'subcategorias': ['Cashback', 'Férias', 'PLR', 'Rendimentos', 'Salario', 'Terceiros']
                        }
                    ]
                    salvar_dados(dados)
                
                # Se não tem contas, criar padrão
                if 'contas' not in dados or len(dados['contas']) == 0:
                    dados['contas'] = [
                        {
                            'id': 1,
                            'bandeira': 'Bradesco',
                            'tipo': 'CORRENTE',
                            'saldo_manual': 100.00,
                            'criada_em': datetime.now().isoformat()
                        },
                        {
                            'id': 2,
                            'bandeira': 'Carteira',
                            'tipo': 'CORRENTE',
                            'saldo_manual': 50.00,
                            'criada_em': datetime.now().isoformat()
                        },
                        {
                            'id': 3,
                            'bandeira': 'C6 Bank',
                            'tipo': 'CORRENTE',
                            'saldo_manual': 75.50,
                            'criada_em': datetime.now().isoformat()
                        },
                        {
                            'id': 4,
                            'bandeira': 'Santander',
                            'tipo': 'CORRENTE',
                            'saldo_manual': 25.00,
                            'criada_em': datetime.now().isoformat()
                        },
                        {
                            'id': 5,
                            'bandeira': 'nuBank',
                            'tipo': 'CARTAO',
                            'saldo_manual': 15.59,
                            'criada_em': datetime.now().isoformat()
                        },
                        {
                            'id': 6,
                            'bandeira': 'Inter',
                            'tipo': 'CARTAO',
                            'saldo_manual': 7.00,
                            'criada_em': datetime.now().isoformat()
                        }
                    ]
                    salvar_dados(dados)
                
                return dados
        except:
            pass
    
    return {
        'contas': [
            {
                'id': 1,
                'bandeira': 'Bradesco',
                'tipo': 'CORRENTE',
                'saldo_manual': 100.00,
                'criada_em': datetime.now().isoformat()
            },
            {
                'id': 2,
                'bandeira': 'Carteira',
                'tipo': 'CORRENTE',
                'saldo_manual': 50.00,
                'criada_em': datetime.now().isoformat()
            },
            {
                'id': 3,
                'bandeira': 'C6 Bank',
                'tipo': 'CORRENTE',
                'saldo_manual': 75.50,
                'criada_em': datetime.now().isoformat()
            },
            {
                'id': 4,
                'bandeira': 'Santander',
                'tipo': 'CORRENTE',
                'saldo_manual': 25.00,
                'criada_em': datetime.now().isoformat()
            },
            {
                'id': 5,
                'bandeira': 'nuBank',
                'tipo': 'CARTAO',
                'saldo_manual': 15.59,
                'criada_em': datetime.now().isoformat()
            },
            {
                'id': 6,
                'bandeira': 'Inter',
                'tipo': 'CARTAO',
                'saldo_manual': 7.00,
                'criada_em': datetime.now().isoformat()
            }
        ],
        'transacoes': [],
        'cartoes': [],
        'faturas': [],
        'categorias': [
            {
                'id': 1,
                'nome': 'Animal_Estimação',
                'tipo': 'DESPESA',
                'subcategorias': ['Creche', 'Pet Shop', 'Ração', 'Remédio', 'Veterinario']
            },
            {
                'id': 2,
                'nome': 'Alimentação',
                'tipo': 'DESPESA',
                'subcategorias': ['Almoco Trabalho', 'Doces', 'Feira', 'Lanche', 'paes', 'Pastel', 'Pizza', 'Restaurante', 'Supermercado']
            },
            {
                'id': 3,
                'nome': 'Casa',
                'tipo': 'DESPESA',
                'subcategorias': ['Brastemp', 'Celular', 'Cloud', 'Comgas', 'Complemento Mensal', 'condominio', 'Eletro Eletronicos', 'Energia Eletrica', 'Faxineira', 'Internet', 'IPTU', 'Manutencao', 'Seguro']
            },
            {
                'id': 4,
                'nome': 'Cuidados_Pessoais',
                'tipo': 'DESPESA',
                'subcategorias': ['Barba', 'Cabelo', 'Cosmeticos', 'Perfume']
            },
            {
                'id': 5,
                'nome': 'Educação',
                'tipo': 'DESPESA',
                'subcategorias': ['Curso', 'Formatura', 'Lanche', 'material escolar', 'Mensalidade', 'Transporte', 'Vestibular']
            },
            {
                'id': 6,
                'nome': 'Eletronicos',
                'tipo': 'DESPESA',
                'subcategorias': ['App', 'Audio / Video', 'pilhas']
            },
            {
                'id': 7,
                'nome': 'Emprestimos',
                'tipo': 'DESPESA',
                'subcategorias': []
            },
            {
                'id': 8,
                'nome': 'Taxas_Impostos',
                'tipo': 'DESPESA',
                'subcategorias': ['Anuidade', 'IOF']
            },
            {
                'id': 9,
                'nome': 'Investimento',
                'tipo': 'DESPESA',
                'subcategorias': ['Açoes', 'Capitalização', 'CDB', 'Previdencia Privada', 'Renda Fixa', 'Tesouro Direto', 'Dólar']
            },
            {
                'id': 10,
                'nome': 'Jogos',
                'tipo': 'DESPESA',
                'subcategorias': ['Aposta', 'Loteria', 'Rifa']
            },
            {
                'id': 11,
                'nome': 'Lazer',
                'tipo': 'DESPESA',
                'subcategorias': ['Academia', 'Balada', 'Bebida', 'Churrasco', 'Cinema', 'Doces', 'Festa', 'Formatura', 'Futebol', 'Show', 'Streaming', 'Video game']
            },
            {
                'id': 12,
                'nome': 'Mesada',
                'tipo': 'DESPESA',
                'subcategorias': ['Gustavo', 'Leonardo']
            },
            {
                'id': 13,
                'nome': 'Outros',
                'tipo': 'DESPESA',
                'subcategorias': ['Não sei', 'Terceiros']
            },
            {
                'id': 14,
                'nome': 'Presentes',
                'tipo': 'DESPESA',
                'subcategorias': ['Aniversario', 'Claudia', 'Dia das maes', 'Dia dos pais', 'Eduardo', 'Gustavo', 'Leonardo', 'Natal', 'Outros']
            },
            {
                'id': 15,
                'nome': 'Saude',
                'tipo': 'DESPESA',
                'subcategorias': ['Convenio medico', 'dentista', 'estacionamento', 'farmacia', 'Oculos', 'Suplementacao']
            },
            {
                'id': 16,
                'nome': 'Transporte',
                'tipo': 'DESPESA',
                'subcategorias': ['Acessorios Carro', 'Aquisicao Veiculo', 'Combustivel', 'estacionamento', 'garagem', 'IPVA / Licenciamento', 'Lavagem', 'manutencao', 'multas', 'pedagio', 'Seguro', 'Transp Publico', 'Uber']
            },
            {
                'id': 17,
                'nome': 'Vestuario',
                'tipo': 'DESPESA',
                'subcategorias': ['Claudia', 'Eduardo', 'Familia', 'Gustavo', 'Leonardo']
            },
            {
                'id': 18,
                'nome': 'Viagem',
                'tipo': 'DESPESA',
                'subcategorias': ['Alimentacao', 'Animal Estimacao', 'Carro', 'Combustivel', 'Dolar', 'Farmacia', 'Hospedagem', 'Passagem', 'Passeio', 'pedagios', 'pontos / milhas', 'Passaporte', 'Presentes', 'Seguro', 'salaVip', 'Visto']
            },
            {
                'id': 19,
                'nome': 'Depositos',
                'tipo': 'RECEITA',
                'subcategorias': ['Cashback', 'Férias', 'PLR', 'Rendimentos', 'Salario', 'Terceiros']
            }
        ]
    }

def salvar_dados(dados):
    """Salva dados no arquivo"""
    with open(DADOS_FILE, 'w') as f:
        json.dump(dados, f, indent=2)

# ============= ROTAS =============

@app.route('/')
def index():
    """Página principal"""
    response = make_response(render_template('simples.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/dashboard')
def dashboard():
    """Página dashboard com gráficos"""
    return render_template('dashboard.html')

# ========== API =========

@app.route('/api/status', methods=['GET'])
def api_status():
    """Testa se API está respondendo"""
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()})

@app.route('/api/contas', methods=['GET'])
def api_contas():
    """Lista contas"""
    dados = carregar_dados()
    return jsonify(dados['contas'])

@app.route('/api/contas', methods=['POST'])
def api_criar_conta():
    """Cria nova conta"""
    try:
        dados = carregar_dados()
        nova_conta = {
            'id': len(dados['contas']) + 1,
            'bandeira': request.json['bandeira'],
            'tipo': request.json['tipo'],
            'saldo_manual': float(request.json.get('saldo_manual', 0)),
            'criada_em': datetime.now().isoformat()
        }
        dados['contas'].append(nova_conta)
        salvar_dados(dados)
        return jsonify(nova_conta), 201
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/transacoes', methods=['GET'])
def api_transacoes():
    """Lista transações"""
    dados = carregar_dados()
    return jsonify(dados['transacoes'])

@app.route('/api/transacoes', methods=['POST'])
def api_criar_transacao():
    """Cria nova transação com todos os detalhes. Se for crédito, gera/atualiza fatura automaticamente"""
    try:
        dados = carregar_dados()
        
        cartao_id = request.json.get('cartao_id', None)
        tipo_conta = request.json.get('tipo_conta', '')
        total_parcelas = request.json.get('total_parcelas', 1)
        eh_recorrente = request.json.get('eh_recorrente', False)
        qtd_repeticoes = request.json.get('qtd_repeticoes', 0)
        
        # Se é recorrente (repetir por X meses), cria múltiplas transações
        if eh_recorrente and qtd_repeticoes > 0:
            data_primeira_recorrencia = request.json.get('data_vencimento', '')
            
            # Transação base para criar as recorrências
            transacao_base = {
                'tipo_conta': tipo_conta,
                'nome_conta': request.json.get('nome_conta', ''),
                'cartao_id': cartao_id,
                'tipo': request.json['tipo'],
                'valor': float(request.json['valor']),
                'status': request.json.get('status', 'PREVISTO'),
                'categoria': request.json['categoria'],
                'subcategoria': request.json.get('subcategoria', ''),
                'tipo_custo': request.json.get('tipo_custo', 'VARIÁVEL'),
                'observacoes': request.json.get('observacoes', ''),
                'descricao': request.json.get('descricao', '')
            }
            
            # Cria as transações recorrentes
            ids_criadas = criar_transacao_recorrente(dados, transacao_base, qtd_repeticoes, data_primeira_recorrencia)
            
            # Gerar/atualizar faturas para cada recorrência (se for crédito)
            if tipo_conta == 'CREDITO' and cartao_id:
                data_primeira = datetime.strptime(data_primeira_recorrencia, '%Y-%m-%d')
                for i in range(qtd_repeticoes):
                    data_recorrencia = data_primeira + relativedelta(months=i)
                    gerar_ou_atualizar_fatura(dados, cartao_id, data_recorrencia.strftime('%Y-%m-%d'))
            
            salvar_dados(dados)
            
            # Retorna info sobre o gasto recorrente
            return jsonify({
                'ids_criadas': ids_criadas,
                'total_repeticoes': qtd_repeticoes,
                'valor': float(request.json['valor']),
                'mensagem': f'✅ Gasto recorrente criado para os próximos {qtd_repeticoes} meses!'
            }), 201
        
        # Se é parcelada (total_parcelas > 1), cria múltiplas transações
        if total_parcelas > 1:
            data_primeira_parcela = request.json.get('data_vencimento', '')
            
            # Transação base para criar as parcelas
            transacao_base = {
                'tipo_conta': tipo_conta,
                'nome_conta': request.json.get('nome_conta', ''),
                'cartao_id': cartao_id,
                'tipo': request.json['tipo'],
                'valor': float(request.json['valor']),
                'status': request.json.get('status', 'PREVISTO'),
                'categoria': request.json['categoria'],
                'subcategoria': request.json.get('subcategoria', ''),
                'tipo_custo': request.json.get('tipo_custo', 'VARIÁVEL'),
                'observacoes': request.json.get('observacoes', ''),
                'descricao': request.json.get('descricao', '')
            }
            
            # Cria as transações parceladas
            ids_criadas = criar_transacao_parcelada(dados, transacao_base, total_parcelas, data_primeira_parcela)
            
            # Gerar/atualizar faturas para cada parcela
            if tipo_conta == 'CREDITO' and cartao_id:
                # Começando da primeira parcela
                data_primeira = datetime.strptime(data_primeira_parcela, '%Y-%m-%d')
                for i in range(total_parcelas):
                    data_parcela = data_primeira + relativedelta(months=i)
                    gerar_ou_atualizar_fatura(dados, cartao_id, data_parcela.strftime('%Y-%m-%d'))
            
            salvar_dados(dados)
            
            # Retorna info sobre a compra parcelada
            return jsonify({
                'ids_criadas': ids_criadas,
                'total_parcelas': total_parcelas,
                'valor_total': float(request.json['valor']),
                'valor_parcela': round(float(request.json['valor']) / total_parcelas, 2),
                'mensagem': f'✅ Compra parcelada em {total_parcelas}x criada com sucesso!'
            }), 201
        
        # Caso de transação simples (não parcelada)
        # Validar tipo de transação
        tipo_transacao = request.json['tipo']
        if tipo_transacao not in ['RECEITA', 'DESPESA', 'TRANSFERENCIA']:
            return jsonify({'erro': 'Tipo deve ser RECEITA, DESPESA ou TRANSFERENCIA'}), 400
        
        # Validar TRANSFERENCIA
        if tipo_transacao == 'TRANSFERENCIA':
            conta_origem_id = request.json.get('conta_origem_id')
            conta_destino_id = request.json.get('conta_destino_id')
            
            if not conta_origem_id or not conta_destino_id:
                return jsonify({'erro': 'TRANSFERENCIA requer conta_origem_id e conta_destino_id'}), 400
            
            if conta_origem_id == conta_destino_id:
                return jsonify({'erro': 'Conta origem e destino não podem ser iguais'}), 400
            
            # Validar que ambas as contas existem e não são cartão de crédito
            conta_origem = next((c for c in dados['contas'] if c['id'] == conta_origem_id), None)
            conta_destino = next((c for c in dados['contas'] if c['id'] == conta_destino_id), None)
            
            if not conta_origem or not conta_destino:
                return jsonify({'erro': 'Uma ou ambas as contas não foram encontradas'}), 404
            
            if conta_origem['tipo'] == 'CARTAO' or conta_destino['tipo'] == 'CARTAO':
                return jsonify({'erro': 'Transferência não permitida com cartão de crédito'}), 400
        
        nova_transacao = {
            'id': max([t['id'] for t in dados['transacoes']], default=0) + 1,
            'data_vencimento': request.json.get('data_vencimento', ''),
            'data_transacao': request.json.get('data_transacao', ''),
            'mes_pagamento': request.json.get('mes_pagamento', ''),
            'tipo_conta': tipo_conta,
            'nome_conta': request.json.get('nome_conta', ''),
            'cartao_id': cartao_id,
            'tipo': tipo_transacao,
            'valor': float(request.json['valor']),
            'parcela': request.json.get('parcela', ''),
            'status': request.json.get('status', 'PREVISTO'),
            'categoria': request.json['categoria'] if tipo_transacao != 'TRANSFERENCIA' else 'Transferência',
            'subcategoria': request.json.get('subcategoria', ''),
            'tipo_custo': request.json.get('tipo_custo', 'VARIÁVEL'),
            'observacoes': request.json.get('observacoes', ''),
            'descricao': request.json.get('descricao', ''),
            'conta_origem_id': request.json.get('conta_origem_id'),
            'conta_destino_id': request.json.get('conta_destino_id'),
            'efetuada': False,
            'criada_em': datetime.now().isoformat()
        }
        dados['transacoes'].append(nova_transacao)
        
        # Se é transação de crédito, gerar/atualizar fatura automaticamente
        if tipo_conta == 'CREDITO' and cartao_id:
            gerar_ou_atualizar_fatura(dados, cartao_id, request.json.get('data_vencimento', ''))
        
        salvar_dados(dados)
        return jsonify(nova_transacao), 201
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

def gerar_ou_atualizar_fatura(dados, cartao_id, data_vencimento):
    """Gera ou atualiza fatura automaticamente baseado nas transações do cartão"""
    try:
        # Extrair mês/ano da data de vencimento
        data_obj = datetime.strptime(data_vencimento, '%Y-%m-%d')
        mes_ano = f"{data_obj.month:02d}/{data_obj.year}"
        
        # Procurar fatura existente aberta para este cartão e mês
        fatura_existente = None
        for f in dados['faturas']:
            if f['cartao_id'] == cartao_id and f['mes'] == mes_ano and f['status'] == 'aberta':
                fatura_existente = f
                break
        
        # Calcular saldo total de transações para este cartão neste mês
        saldo_total = 0
        for t in dados['transacoes']:
            if t.get('cartao_id') == cartao_id and t['mes_pagamento'] == mes_ano:
                if t['tipo'] == 'DESPESA':
                    saldo_total += float(t['valor'])
                elif t['tipo'] == 'RECEITA':
                    saldo_total -= float(t['valor'])
        
        if fatura_existente:
            # Atualizar saldo da fatura existente
            fatura_existente['saldo'] = saldo_total
        else:
            # Criar nova fatura
            cartao = next((c for c in dados['cartoes'] if c['id'] == cartao_id), None)
            if cartao:
                nova_fatura = {
                    'id': max([f['id'] for f in dados['faturas']], default=0) + 1,
                    'cartao_id': cartao_id,
                    'mes': mes_ano,
                    'data_fechamento': cartao.get('data_fechamento', ''),
                    'data_vencimento': cartao.get('data_vencimento', ''),
                    'status': 'aberta',
                    'saldo': saldo_total,
                    'criada_em': datetime.now().isoformat()
                }
                dados['faturas'].append(nova_fatura)
    except Exception as e:
        print(f"Erro ao gerar fatura: {e}")

def atualizar_saldo_fatura(dados, cartao_id, mes_ano):
    """Recalcula saldo da fatura baseado nas transações"""
    try:
        if not mes_ano:
            return
        
        # Procurar fatura aberta para este cartão e mês
        fatura = None
        for f in dados['faturas']:
            if f['cartao_id'] == cartao_id and f['mes'] == mes_ano and f['status'] == 'aberta':
                fatura = f
                break
        
        if not fatura:
            return
        
        # Recalcular saldo total de transações para este cartão neste mês
        saldo_total = 0
        for t in dados['transacoes']:
            if t.get('cartao_id') == cartao_id and t['mes_pagamento'] == mes_ano:
                if t['tipo'] == 'DESPESA':
                    saldo_total += float(t['valor'])
                elif t['tipo'] == 'RECEITA':
                    saldo_total -= float(t['valor'])
        
        fatura['saldo'] = saldo_total
    except Exception as e:
        print(f"Erro ao atualizar saldo fatura: {e}")

def criar_transacao_parcelada(dados, transacao_base, total_parcelas, data_primeira_parcela):
    """
    Cria N transações (uma por parcela) a partir de uma transação base
    
    Entrada:
    - transacao_base: dict com dados da compra
    - total_parcelas: int (2 até N...)
    - data_primeira_parcela: str no formato 'YYYY-MM-DD'
    
    Retorna: lista de IDs das transações criadas
    """
    try:
        compra_id = f"COMPRA_{str(uuid.uuid4())[:8]}"
        valor_total = float(transacao_base['valor'])
        valor_parcela = valor_total / total_parcelas
        
        data_primeira = datetime.strptime(data_primeira_parcela, '%Y-%m-%d')
        
        ids_criadas = []
        
        for num_parcela in range(1, total_parcelas + 1):
            # Calcular data e mês/ano desta parcela
            data_parcela = data_primeira + relativedelta(months=num_parcela - 1)
            mes_ano = f"{data_parcela.month:02d}/{data_parcela.year}"
            data_str = data_parcela.strftime('%Y-%m-%d')
            
            transacao = {
                'id': max([t['id'] for t in dados['transacoes']], default=0) + 1,
                'data_vencimento': data_str,
                'data_transacao': datetime.now().strftime('%Y-%m-%d'),
                'mes_pagamento': mes_ano,
                'tipo_conta': transacao_base['tipo_conta'],
                'nome_conta': transacao_base.get('nome_conta', ''),
                'cartao_id': transacao_base.get('cartao_id', None),
                'tipo': transacao_base['tipo'],
                'valor': round(valor_parcela, 2),
                'valor_total': valor_total,
                'compra_id': compra_id,
                'num_parcela': num_parcela,
                'total_parcelas': total_parcelas,
                'parcela': f"{num_parcela}/{total_parcelas}",
                'status': transacao_base.get('status', 'PREVISTO'),
                'categoria': transacao_base['categoria'],
                'subcategoria': transacao_base.get('subcategoria', ''),
                'tipo_custo': transacao_base.get('tipo_custo', 'VARIÁVEL'),
                'observacoes': transacao_base.get('observacoes', ''),
                'descricao': f"{transacao_base.get('descricao', '')} ({num_parcela}/{total_parcelas})",
                'efetuada': False,
                'criada_em': datetime.now().isoformat()
            }
            
            dados['transacoes'].append(transacao)
            ids_criadas.append(transacao['id'])
        
        return ids_criadas
    except Exception as e:
        print(f"Erro ao criar transação parcelada: {e}")
        return []

def criar_transacao_recorrente(dados, transacao_base, qtd_repeticoes, data_primeira_recorrencia):
    """
    Cria N transações (uma por mês) com a mesma descrição/categoria
    
    Entrada:
    - transacao_base: dict com dados do gasto recorrente
    - qtd_repeticoes: int (quantos meses repetir: 1-60)
    - data_primeira_recorrencia: str no formato 'YYYY-MM-DD'
    
    Retorna: lista de IDs das transações criadas
    """
    try:
        recorrencia_id = f"REC_{str(uuid.uuid4())[:8]}"
        data_primeira = datetime.strptime(data_primeira_recorrencia, '%Y-%m-%d')
        
        ids_criadas = []
        
        for num_recorrencia in range(1, qtd_repeticoes + 1):
            # Calcular data desta recorrência (mesmo dia, próximo mês)
            data_recorrencia = data_primeira + relativedelta(months=num_recorrencia - 1)
            mes_ano = f"{data_recorrencia.month:02d}/{data_recorrencia.year}"
            data_str = data_recorrencia.strftime('%Y-%m-%d')
            
            transacao = {
                'id': max([t['id'] for t in dados['transacoes']], default=0) + 1,
                'data_vencimento': data_str,
                'data_transacao': datetime.now().strftime('%Y-%m-%d'),
                'mes_pagamento': mes_ano,
                'tipo_conta': transacao_base['tipo_conta'],
                'nome_conta': transacao_base.get('nome_conta', ''),
                'cartao_id': transacao_base.get('cartao_id', None),
                'tipo': transacao_base['tipo'],
                'valor': round(float(transacao_base['valor']), 2),
                'status': transacao_base.get('status', 'PREVISTO'),
                'categoria': transacao_base['categoria'],
                'subcategoria': transacao_base.get('subcategoria', ''),
                'tipo_custo': transacao_base.get('tipo_custo', 'VARIÁVEL'),
                'observacoes': transacao_base.get('observacoes', ''),
                'descricao': transacao_base.get('descricao', ''),
                'eh_recorrente': True,
                'recorrencia_id': recorrencia_id,
                'num_recorrencia': num_recorrencia,
                'total_recorrencias': qtd_repeticoes,
                'efetuada': False,
                'criada_em': datetime.now().isoformat()
            }
            
            dados['transacoes'].append(transacao)
            ids_criadas.append(transacao['id'])
        
        return ids_criadas
    except Exception as e:
        print(f"Erro ao criar transação recorrente: {e}")
        return []

@app.route('/api/contas/<int:conta_id>', methods=['PUT'])
def api_atualizar_conta(conta_id):
    """Atualiza dados de uma conta (nome, tipo, saldo, etc)"""
    try:
        dados = carregar_dados()
        
        # Procura a conta
        conta = next((c for c in dados['contas'] if c['id'] == conta_id), None)
        if not conta:
            print(f"❌ Conta não encontrada: id={conta_id}, contas existentes: {[c['id'] for c in dados['contas']]}")
            return jsonify({'erro': f'Conta {conta_id} não encontrada'}), 404
        
        # Log do que está vindo no request
        print(f"📝 Atualizando conta {conta_id}:")
        print(f"   Request JSON: {request.json}")
        
        # Atualiza bandeira (nome) se fornecido
        if 'bandeira' in request.json:
            conta['bandeira'] = request.json['bandeira']
            print(f"   ✓ Bandeira atualizada: {conta['bandeira']}")
        
        # Atualiza tipo se fornecido (CRÍTICO!)
        if 'tipo' in request.json:
            conta['tipo'] = request.json['tipo']
            print(f"   ✓ Tipo atualizado: {conta['tipo']}")
        else:
            print(f"   ⚠️  Tipo NÃO fornecido no request!")
        
        # Atualiza saldo se fornecido
        if 'saldo_manual' in request.json:
            conta['saldo_manual'] = float(request.json['saldo_manual'])
            print(f"   ✓ Saldo atualizado: {conta['saldo_manual']}")
        
        print(f"   Conta após atualização: {conta}")
        salvar_dados(dados)
        print(f"   ✅ Conta salva com sucesso")
        return jsonify(conta), 200
    except Exception as e:
        print(f"❌ Erro ao atualizar conta: {str(e)}")
        return jsonify({'erro': str(e)}), 400

@app.route('/api/transacoes/<int:transacao_id>/efetivar', methods=['PUT'])
def api_efetivar_transacao(transacao_id):
    """Efetiva uma transação sensibilizando o saldo se data for passada/presente"""
    try:
        dados = carregar_dados()
        
        # Procura a transação
        transacao = next((t for t in dados['transacoes'] if t['id'] == transacao_id), None)
        if not transacao:
            return jsonify({'erro': 'Transação não encontrada'}), 404
        
        # Já foi efetuada?
        if transacao.get('efetuada', False):
            return jsonify({'erro': 'Transação já foi efetuada'}), 400
        
        # Verifica a data
        data_vencimento = transacao.get('data_vencimento', '')
        if not data_vencimento:
            return jsonify({'erro': 'Transação sem data de vencimento'}), 400
        
        try:
            data_trans = datetime.strptime(data_vencimento, '%Y-%m-%d').date()
            data_hoje = datetime.now().date()
            
            tipo_conta = transacao.get('tipo_conta', 'CORRENTE')
            
            # Se for TRANSFERENCIA: debitar origem e creditar destino
            if transacao['tipo'] == 'TRANSFERENCIA':
                conta_origem_id = transacao.get('conta_origem_id')
                conta_destino_id = transacao.get('conta_destino_id')
                
                if not conta_origem_id or not conta_destino_id:
                    return jsonify({'erro': 'Transferência com contas inválidas'}), 400
                
                conta_origem = next((c for c in dados['contas'] if c['id'] == conta_origem_id), None)
                conta_destino = next((c for c in dados['contas'] if c['id'] == conta_destino_id), None)
                
                if not conta_origem or not conta_destino:
                    return jsonify({'erro': 'Uma ou ambas as contas não foram encontradas'}), 404
                
                # Permite saldo negativo em contas corrente - comentado
                # if conta_origem['saldo_manual'] < transacao['valor']:
                #     return jsonify({'erro': f'Saldo insuficiente na conta {conta_origem["bandeira"]}'}), 400
                
                # Efetiva a transferência
                conta_origem['saldo_manual'] -= transacao['valor']
                conta_destino['saldo_manual'] += transacao['valor']
                
                transacao['efetuada'] = True
                transacao['status'] = 'PAGO'
                
                print(f'✅ Transferência efetuada: {conta_origem["bandeira"]} → {conta_destino["bandeira"]}: R$ {transacao["valor"]}')
                
                salvar_dados(dados)
                return jsonify({
                    'transacao': transacao,
                    'conta_origem': conta_origem,
                    'conta_destino': conta_destino,
                    'mensagem': f'Transferência de R$ {transacao["valor"]} realizada com sucesso'
                }), 200
            
            # Se for CRÉDITO: apenas marcar como efetuada (reconhecimento)
            if tipo_conta == 'CREDITO':
                print(f'✅ Efetuando transação CRÉDITO ID {transacao_id} - apenas reconhecimento')
                transacao['efetuada'] = True
                # Sinaliza como AGENDADO
                transacao['status'] = 'AGENDADO'
                
                salvar_dados(dados)
                return jsonify({
                    'transacao': transacao,
                    'mensagem': '✅ Transação de crédito reconhecida. Fatura e saldo projetado atualizados.'
                }), 200
            
            # Se for CORRENTE: sensibilizar o saldo da conta
            elif tipo_conta == 'CORRENTE':
                # Busca a conta pelo nome
                nome_conta = transacao.get('nome_conta', '')
                conta = next((c for c in dados['contas'] if c['bandeira'] == nome_conta), None)
                
                if not conta:
                    return jsonify({'erro': 'Conta referenciada não encontrada'}), 404
                
                # Sensibiliza o saldo
                if transacao['tipo'] == 'DESPESA':
                    conta['saldo_manual'] -= transacao['valor']
                elif transacao['tipo'] == 'RECEITA':
                    conta['saldo_manual'] += transacao['valor']
                
                # Marca como efetuada
                transacao['efetuada'] = True
                
                # Muda o status automaticamente baseado no tipo
                if transacao['tipo'] == 'DESPESA':
                    transacao['status'] = 'PAGO'
                elif transacao['tipo'] == 'RECEITA':
                    transacao['status'] = 'RECEBIDO'
                
                print(f'✅ Efetuando transação CORRENTE ID {transacao_id} - sensibilizando saldo de {nome_conta}')
                
                salvar_dados(dados)
                return jsonify({
                    'transacao': transacao,
                    'conta': conta,
                    'mensagem': 'Transação efetuada com sucesso'
                }), 200
            
        except ValueError:
            return jsonify({'erro': 'Formato de data inválido'}), 400
            
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/transacoes/<int:transacao_id>', methods=['DELETE'])
def api_deletar_transacao(transacao_id):
    """Deleta uma transação e reverte saldo se foi efetuada"""
    try:
        dados = carregar_dados()
        
        # Procura a transação
        transacao = next((t for t in dados['transacoes'] if t['id'] == transacao_id), None)
        if not transacao:
            return jsonify({'erro': 'Transação não encontrada'}), 404
        
        # Se foi efetuada, reverte o saldo
        if transacao.get('efetuada', False):
            # Se é transferência, reverter em ambas as contas
            if transacao['tipo'] == 'TRANSFERENCIA':
                conta_origem_id = transacao.get('conta_origem_id')
                conta_destino_id = transacao.get('conta_destino_id')
                
                if conta_origem_id and conta_destino_id:
                    conta_origem = next((c for c in dados['contas'] if c['id'] == conta_origem_id), None)
                    conta_destino = next((c for c in dados['contas'] if c['id'] == conta_destino_id), None)
                    
                    if conta_origem and conta_destino:
                        # Reverte a transferência (inverte os valores)
                        conta_origem['saldo_manual'] += transacao['valor']  # Adiciona de volta
                        conta_destino['saldo_manual'] -= transacao['valor']  # Remove
            else:
                # Reversão para RECEITA/DESPESA
                nome_conta = transacao.get('nome_conta', '')
                conta = next((c for c in dados['contas'] if c['bandeira'] == nome_conta), None)
                
                if conta:
                    # Reverte a operação (inverte a lógica)
                    if transacao['tipo'] == 'DESPESA':
                        conta['saldo_manual'] += transacao['valor']  # Adiciona de volta (era despesa)
                    elif transacao['tipo'] == 'RECEITA':
                        conta['saldo_manual'] -= transacao['valor']  # Remove (era receita)
        
        # Se é transação de crédito, atualizar fatura
        if transacao.get('tipo_conta') == 'CREDITO' and transacao.get('cartao_id'):
            atualizar_saldo_fatura(dados, transacao.get('cartao_id'), transacao.get('mes_pagamento', ''))
        
        # Remove da lista
        dados['transacoes'] = [t for t in dados['transacoes'] if t['id'] != transacao_id]
        salvar_dados(dados)
        
        return jsonify({
            'mensagem': 'Transação deletada com sucesso',
            'saldo_revertido': transacao.get('efetuada', False)
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/transacoes/<int:transacao_id>', methods=['PUT'])
def api_editar_transacao(transacao_id):
    """Edita uma transação. Se for crédito, atualiza fatura automaticamente"""
    try:
        dados = carregar_dados()
        
        # Procura a transação
        transacao = next((t for t in dados['transacoes'] if t['id'] == transacao_id), None)
        if not transacao:
            return jsonify({'erro': 'Transação não encontrada'}), 404
        
        # Não permite editar se já foi efetuada
        if transacao.get('efetuada', False):
            return jsonify({'erro': 'Não é permitido editar transação já efetuada'}), 400
        
        # Guardar mes_pagamento antigo para atualizar fatura corretamente
        mes_antigo = transacao.get('mes_pagamento', '')
        cartao_id_antigo = transacao.get('cartao_id')
        
        # Atualiza os campos fornecidos
        campos_editaveis = ['data_vencimento', 'data_transacao', 'mes_pagamento', 'tipo_conta', 'nome_conta', 
                           'cartao_id', 'tipo', 'valor', 'parcela', 'status', 'categoria', 'subcategoria', 
                           'tipo_custo', 'observacoes', 'descricao']
        
        for campo in campos_editaveis:
            if campo in request.json:
                if campo == 'valor':
                    transacao[campo] = float(request.json[campo])
                elif campo == 'cartao_id':
                    transacao[campo] = int(request.json[campo]) if request.json[campo] else None
                else:
                    transacao[campo] = request.json[campo]
        
        # Se é transação de crédito, atualizar faturas
        mes_novo = transacao.get('mes_pagamento', '')
        if transacao.get('tipo_conta') == 'CREDITO' and transacao.get('cartao_id'):
            # Atualizar fatura do mês antigo se mudou o mês
            if mes_antigo and mes_antigo != mes_novo and cartao_id_antigo:
                atualizar_saldo_fatura(dados, cartao_id_antigo, mes_antigo)
            # Atualizar fatura do mês novo
            atualizar_saldo_fatura(dados, transacao.get('cartao_id'), mes_novo)
        
        salvar_dados(dados)
        return jsonify({
            'transacao': transacao,
            'mensagem': 'Transação editada com sucesso'
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/resumo', methods=['GET'])
def api_resumo():
    """Resumo financeiro"""
    try:
        dados = carregar_dados()
        
        total_receita = sum(t['valor'] for t in dados['transacoes'] if t['tipo'] == 'RECEITA')
        total_despesa = sum(t['valor'] for t in dados['transacoes'] if t['tipo'] == 'DESPESA')
        saldo_manual = sum(c['saldo_manual'] for c in dados['contas'])
        
        # Calcular o mês de pagamento do período de fechamento corrente
        # Ciclo: começa no penúltimo dia útil de um mês, vai até antepenúltimo dia útil do próximo
        # Se hoje >= 29, já estamos no próximo ciclo (próximo mês de pagamento)
        hoje = datetime.now()
        
        if hoje.day >= 29:
            mes_pagamento_corrente_obj = hoje + relativedelta(months=1)
        else:
            mes_pagamento_corrente_obj = hoje
        
        mes_pagamento_corrente = f"{mes_pagamento_corrente_obj.month:02d}/{mes_pagamento_corrente_obj.year}"
        
        # Calcular receitas pendentes (não efetivadas) do período de fechamento corrente
        receitas_pendentes = sum(
            t['valor'] 
            for t in dados['transacoes'] 
            if t['tipo'] == 'RECEITA' 
            and t.get('status') in ['PREVISTO', 'AGENDADO', 'REC_PENDENTE']
            and t.get('efetuada') == False
            and t.get('mes_pagamento') == mes_pagamento_corrente
        )
        
        # Calcular despesas pendentes (não efetivadas) do período de fechamento corrente
        despesas_pendentes = sum(
            t['valor'] 
            for t in dados['transacoes'] 
            if t['tipo'] == 'DESPESA' 
            and t.get('status') in ['PREVISTO', 'AGENDADO']
            and t.get('efetuada') == False
            and t.get('mes_pagamento') == mes_pagamento_corrente
        )
        
        return jsonify({
            'total_receita': total_receita,
            'total_despesa': total_despesa,
            'receitas_pendentes': receitas_pendentes,
            'despesas_pendentes': despesas_pendentes,
            'saldo_manual': saldo_manual,
            'saldo_liquido': saldo_manual + total_receita - total_despesa,
            'num_contas': len(dados['contas']),
            'num_transacoes': len(dados['transacoes'])
        })
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

# ========== CATEGORIAS ==========

@app.route('/api/categorias', methods=['GET'])
def api_categorias():
    """Lista todas as categorias"""
    dados = carregar_dados()
    return jsonify(dados['categorias'])

@app.route('/api/categorias/tipo/<tipo>', methods=['GET'])
def api_categorias_por_tipo(tipo):
    """Lista categorias por tipo (RECEITA ou DESPESA)"""
    dados = carregar_dados()
    categorias_filtradas = [c for c in dados['categorias'] if c['tipo'] == tipo]
    return jsonify(categorias_filtradas)

@app.route('/api/categorias', methods=['POST'])
def api_criar_categoria():
    """Cria nova categoria"""
    try:
        dados = carregar_dados()
        
        # Valida tipo
        tipo = request.json.get('tipo', '').upper()
        if tipo not in ['RECEITA', 'DESPESA']:
            return jsonify({'erro': 'Tipo deve ser RECEITA ou DESPESA'}), 400
        
        # Verifica se categoria já existe
        if any(c['nome'].lower() == request.json['nome'].lower() for c in dados['categorias']):
            return jsonify({'erro': 'Categoria já existe'}), 400
        
        nova_categoria = {
            'id': max([c['id'] for c in dados['categorias']], default=0) + 1,
            'nome': request.json['nome'],
            'tipo': tipo,
            'subcategorias': []
        }
        
        dados['categorias'].append(nova_categoria)
        salvar_dados(dados)
        return jsonify(nova_categoria), 201
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/categorias/<int:categoria_id>', methods=['PUT'])
def api_atualizar_categoria(categoria_id):
    """Atualiza categoria"""
    try:
        dados = carregar_dados()
        categoria = next((c for c in dados['categorias'] if c['id'] == categoria_id), None)
        
        if not categoria:
            return jsonify({'erro': 'Categoria não encontrada'}), 404
        
        if 'nome' in request.json:
            categoria['nome'] = request.json['nome']
        
        if 'tipo' in request.json:
            tipo = request.json['tipo'].upper()
            if tipo not in ['RECEITA', 'DESPESA']:
                return jsonify({'erro': 'Tipo deve ser RECEITA ou DESPESA'}), 400
            categoria['tipo'] = tipo
        
        salvar_dados(dados)
        return jsonify(categoria), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/categorias/<int:categoria_id>', methods=['DELETE'])
def api_deletar_categoria(categoria_id):
    """Deleta categoria"""
    try:
        dados = carregar_dados()
        categoria = next((c for c in dados['categorias'] if c['id'] == categoria_id), None)
        
        if not categoria:
            return jsonify({'erro': 'Categoria não encontrada'}), 404
        
        dados['categorias'] = [c for c in dados['categorias'] if c['id'] != categoria_id]
        salvar_dados(dados)
        
        return jsonify({'mensagem': 'Categoria deletada com sucesso'}), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

# ========== SUBCATEGORIAS ==========

@app.route('/api/categorias/<int:categoria_id>/subcategorias', methods=['POST'])
def api_criar_subcategoria(categoria_id):
    """Cria nova subcategoria para uma categoria"""
    try:
        dados = carregar_dados()
        categoria = next((c for c in dados['categorias'] if c['id'] == categoria_id), None)
        
        if not categoria:
            return jsonify({'erro': 'Categoria mãe não encontrada'}), 404
        
        nome_subcategoria = request.json.get('nome', '').strip()
        
        # Verifica se subcategoria já existe nesta categoria
        if any(s.lower() == nome_subcategoria.lower() for s in categoria.get('subcategorias', [])):
            return jsonify({'erro': 'Subcategoria já existe nesta categoria'}), 400
        
        if 'subcategorias' not in categoria:
            categoria['subcategorias'] = []
        
        categoria['subcategorias'].append(nome_subcategoria)
        salvar_dados(dados)
        
        return jsonify({
            'categoria_id': categoria_id,
            'subcategoria': nome_subcategoria,
            'mensagem': 'Subcategoria criada com sucesso'
        }), 201
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/categorias/<int:categoria_id>/subcategorias/<path:nome_subcategoria>', methods=['DELETE'])
def api_deletar_subcategoria(categoria_id, nome_subcategoria):
    """Deleta uma subcategoria"""
    try:
        dados = carregar_dados()
        categoria = next((c for c in dados['categorias'] if c['id'] == categoria_id), None)
        
        if not categoria:
            return jsonify({'erro': 'Categoria não encontrada'}), 404
        
        # Decodifica o nome
        nome_subcategoria = urllib.parse.unquote(nome_subcategoria)
        
        if nome_subcategoria not in categoria.get('subcategorias', []):
            return jsonify({'erro': 'Subcategoria não encontrada'}), 404
        
        categoria['subcategorias'].remove(nome_subcategoria)
        salvar_dados(dados)
        
        return jsonify({'mensagem': 'Subcategoria deletada com sucesso'}), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

# ========== CARTÕES DE CRÉDITO ==========

@app.route('/api/cartoes', methods=['GET'])
def api_cartoes():
    """Lista todos os cartões de crédito"""
    dados = carregar_dados()
    return jsonify(dados['cartoes'])

@app.route('/api/cartoes', methods=['POST'])
def api_criar_cartao():
    """Cria novo cartão de crédito"""
    try:
        dados = carregar_dados()
        novo_cartao = {
            'id': max([c['id'] for c in dados['cartoes']], default=0) + 1,
            'bandeira': request.json['bandeira'],
            'ultimos_digitos': request.json.get('ultimos_digitos', ''),
            'limite_credito': float(request.json.get('limite_credito', 0)),
            'data_fechamento': request.json.get('data_fechamento', ''),
            'data_vencimento': request.json.get('data_vencimento', ''),
            'status': 'ATIVO',
            'criado_em': datetime.now().isoformat()
        }
        dados['cartoes'].append(novo_cartao)
        salvar_dados(dados)
        return jsonify(novo_cartao), 201
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/cartoes/<int:cartao_id>', methods=['PUT'])
def api_atualizar_cartao(cartao_id):
    """Atualiza dados de um cartão"""
    try:
        dados = carregar_dados()
        cartao = next((c for c in dados['cartoes'] if c['id'] == cartao_id), None)
        
        if not cartao:
            return jsonify({'erro': 'Cartão não encontrado'}), 404
        
        campos_editaveis = ['bandeira', 'ultimos_digitos', 'limite_credito', 
                           'data_fechamento', 'data_vencimento', 'status']
        
        for campo in campos_editaveis:
            if campo in request.json:
                if campo == 'limite_credito':
                    cartao[campo] = float(request.json[campo])
                else:
                    cartao[campo] = request.json[campo]
        
        salvar_dados(dados)
        return jsonify(cartao), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/cartoes/<int:cartao_id>', methods=['DELETE'])
def api_deletar_cartao(cartao_id):
    """Deleta um cartão"""
    try:
        dados = carregar_dados()
        
        # Verifica se há faturas abertas
        faturas_abertas = [f for f in dados['faturas'] 
                          if f['cartao_id'] == cartao_id and f['status'] == 'aberta']
        
        if faturas_abertas:
            return jsonify({'erro': 'Não é possível deletar cartão com faturas abertas'}), 400
        
        # Remove a atura
        dados['cartoes'] = [c for c in dados['cartoes'] if c['id'] != cartao_id]
        dados['faturas'] = [f for f in dados['faturas'] if f['cartao_id'] != cartao_id]
        
        salvar_dados(dados)
        return jsonify({'mensagem': 'Cartão deletado com sucesso'}), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

# ========== FATURAS ==========

@app.route('/api/cartoes/<int:cartao_id>/faturas', methods=['GET'])
def api_faturas_cartao(cartao_id):
    """Lista faturas de um cartão"""
    dados = carregar_dados()
    
    # Verifica se cartão existe
    cartao = next((c for c in dados['cartoes'] if c['id'] == cartao_id), None)
    if not cartao:
        return jsonify({'erro': 'Cartão não encontrado'}), 404
    
    faturas = [f for f in dados['faturas'] if f['cartao_id'] == cartao_id]
    return jsonify(faturas)

@app.route('/api/faturas', methods=['GET'])
def api_faturas():
    """Lista todas as faturas"""
    dados = carregar_dados()
    return jsonify(dados['faturas'])

@app.route('/api/faturas', methods=['POST'])
def api_criar_fatura():
    """Cria nova fatura"""
    try:
        dados = carregar_dados()
        
        # Verifica se cartão existe
        cartao_id = request.json['cartao_id']
        cartao = next((c for c in dados['cartoes'] if c['id'] == cartao_id), None)
        if not cartao:
            return jsonify({'erro': 'Cartão não encontrado'}), 404
        
        nova_fatura = {
            'id': max([f['id'] for f in dados['faturas']], default=0) + 1,
            'cartao_id': cartao_id,
            'mes': request.json.get('mes', ''),
            'data_fechamento': request.json.get('data_fechamento', ''),
            'data_vencimento': request.json.get('data_vencimento', ''),
            'status': 'aberta',  # Estados: aberta, fechada
            'saldo': float(request.json.get('saldo', 0)),
            'criada_em': datetime.now().isoformat()
        }
        
        dados['faturas'].append(nova_fatura)
        salvar_dados(dados)
        return jsonify(nova_fatura), 201
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/faturas/<int:fatura_id>', methods=['PUT'])
def api_atualizar_fatura(fatura_id):
    """Atualiza uma fatura (pode fechar)"""
    try:
        dados = carregar_dados()
        fatura = next((f for f in dados['faturas'] if f['id'] == fatura_id), None)
        
        if not fatura:
            return jsonify({'erro': 'Fatura não encontrada'}), 404
        
        # Campos editáveis
        campos_editaveis = ['mes', 'data_fechamento', 'data_vencimento', 'status', 'saldo']
        
        for campo in campos_editaveis:
            if campo in request.json:
                if campo == 'saldo':
                    fatura[campo] = float(request.json[campo])
                else:
                    fatura[campo] = request.json[campo]
        
        salvar_dados(dados)
        return jsonify(fatura), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/faturas/<int:fatura_id>', methods=['DELETE'])
def api_deletar_fatura(fatura_id):
    """Deleta uma fatura"""
    try:
        dados = carregar_dados()
        fatura = next((f for f in dados['faturas'] if f['id'] == fatura_id), None)
        
        if not fatura:
            return jsonify({'erro': 'Fatura não encontrada'}), 404
        
        # Não permite deletar fatura fechada
        if fatura['status'] == 'fechada':
            return jsonify({'erro': 'Não é possível deletar fatura fechada'}), 400
        
        dados['faturas'] = [f for f in dados['faturas'] if f['id'] != fatura_id]
        salvar_dados(dados)
        
        return jsonify({'mensagem': 'Fatura deletada com sucesso'}), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/faturas/regenerar', methods=['POST'])
def api_regenerar_faturas():
    """Regenera todas as faturas baseado nas transações de crédito"""
    try:
        dados = carregar_dados()
        
        # Limpar faturas existentes
        dados['faturas'] = []
        
        # Para cada transação de CREDITO, gerar fatura
        transacoes_credito = [t for t in dados['transacoes'] if t.get('tipo_conta') == 'CREDITO']
        
        for transacao in transacoes_credito:
            gerar_ou_atualizar_fatura(dados, transacao['cartao_id'], transacao['data_vencimento'])
        
        salvar_dados(dados)
        
        return jsonify({
            'mensagem': f'✅ {len(dados["faturas"])} faturas regeneradas com sucesso',
            'faturas': dados['faturas']
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/faturas/<int:fatura_id>/pagar', methods=['POST'])
def api_pagar_fatura(fatura_id):
    """Paga uma fatura debitando de uma conta corrente"""
    try:
        dados = carregar_dados()
        
        fatura = next((f for f in dados['faturas'] if f['id'] == fatura_id), None)
        if not fatura:
            return jsonify({'erro': 'Fatura não encontrada'}), 404
        
        conta_id = request.json.get('conta_id')
        valor_pago = float(request.json.get('valor_pago', 0))
        
        if valor_pago <= 0:
            return jsonify({'erro': 'Valor deve ser maior que zero'}), 400
        
        # Encontrar conta corrente
        conta = next((c for c in dados['contas'] if c['id'] == conta_id and c['tipo'] == 'CORRENTE'), None)
        if not conta:
            return jsonify({'erro': 'Conta corrente não encontrada'}), 404
        
        # Debitar da conta
        conta['saldo_manual'] = float(conta['saldo_manual']) - valor_pago
        
        # Subtrair valor da fatura
        fatura['saldo'] = float(fatura['saldo']) - valor_pago
        
        # Se fatura zerada ou negativa, marcar como paga
        if fatura['saldo'] <= 0:
            fatura['saldo'] = 0
            fatura['status'] = 'paga'
            
            # Atualizar status de todas as transações que compõem a fatura para PAGO
            for t in dados['transacoes']:
                if t.get('cartao_id') == fatura['cartao_id'] and t.get('mes_pagamento') == fatura['mes']:
                    t['status'] = 'PAGO'
        
        salvar_dados(dados)
        
        return jsonify({
            'mensagem': f'✅ Pagamento de R$ {valor_pago:.2f} realizado!',
            'fatura': fatura,
            'conta': conta
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

# ===== ENDPOINTS DE IMPORT/EXPORT =====

@app.route('/api/limpar-dados', methods=['POST'])
def limpar_dados():
    """Limpa todos os dados da aplicação"""
    try:
        dados_vazios = {
            'contas': [],
            'transacoes': [],
            'cartoes': [],
            'categorias': [],
            'faturas': []
        }
        salvar_dados(dados_vazios)
        return jsonify({'mensagem': '✅ Todos os dados foram removidos!'}), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/importar-excel', methods=['POST'])
def importar_excel():
    """Importa dados do arquivo base.xlsx
    
    Mapeamento de colunas:
    - Tipo Conta (Excel) → tipo_conta na transação (CORRENTE ou CREDITO)
    - Cartao (Excel) → nome_conta na transação (identificador real da conta)
    """
    try:
        import openpyxl
        from datetime import datetime as dt
        
        arquivo = '/Users/eduardomoretti/Downloads/vscode/base.xlsx'
        if not os.path.exists(arquivo):
            return jsonify({'erro': 'Arquivo base.xlsx não encontrado'}), 404
        
        # Carrega planilha
        wb = openpyxl.load_workbook(arquivo)
        ws = wb['Planilha1']
        
        # Fetch dados atuais (para manter contas/cartões existentes se houver)
        dados = carregar_dados()
        
        # Se vazio, inicializa
        if not dados.get('transacoes'):
            dados['transacoes'] = []
        if not dados.get('contas'):
            dados['contas'] = []
        if not dados.get('cartoes'):
            dados['cartoes'] = []
        if not dados.get('categorias'):
            dados['categorias'] = []
        
        # Lê cabeçalho
        cabecalho = {}
        for i, cell in enumerate(ws[1], 1):
            if cell.value:
                cabecalho[cell.value.lower() if isinstance(cell.value, str) else str(cell.value).lower()] = i
        
        # Processa linhas
        contas_criadas = set()
        cartoes_criados = set()
        transacoes_importadas = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
            try:
                # Extrai dados da linha
                data_transacao = row[cabecalho.get('data transacao', 1) - 1]
                data_vcto = row[cabecalho.get('data vcto', 2) - 1]
                tipo_transacao = row[cabecalho.get('tipo transacao', 4) - 1]
                tipo_conta_excel = row[cabecalho.get('tipo conta', 5) - 1]  # CORRENTE ou CREDITO
                cartao_nome_excel = row[cabecalho.get('cartao', 6) - 1]  # Nome real da conta
                valor = row[cabecalho.get('valor', 7) - 1]
                parcela_info = row[cabecalho.get('parcela', 8) - 1]
                descricao = row[cabecalho.get('onde', 9) - 1]
                status = row[cabecalho.get('status', 10) - 1]
                categoria = row[cabecalho.get('categoria', 11) - 1]
                subcategoria = row[cabecalho.get('sub categoria', 12) - 1]
                tipo_custo = row[cabecalho.get('tipo custo', 13) - 1]
                observacacao = row[cabecalho.get('observação', 14) - 1]
                
                # Validações
                if not data_transacao or not valor or not tipo_transacao:
                    continue
                
                # Converte valor para float
                try:
                    valor = float(valor)
                except:
                    continue
                
                # Processa datas
                if isinstance(data_transacao, dt):
                    data_str = data_transacao.strftime('%Y-%m-%d')
                else:
                    data_str = str(data_transacao)
                
                if isinstance(data_vcto, dt):
                    data_vcto_str = data_vcto.strftime('%Y-%m-%d')
                else:
                    data_vcto_str = str(data_vcto)
                
                # Obtém mês de pagamento
                if isinstance(data_vcto, dt):
                    mes_pgto = data_vcto.strftime('%m/%Y')
                else:
                    mes_pgto = '03/2026'
                
                # NOVO MAPEAMENTO CORRETO:
                # Tipo Conta (Excel) → tipo_conta na aplicação (CORRENTE ou CREDITO)
                # Note: Usar CREDITO (não CARTAO) para consistência com backend
                tipo_conta_normalizado = 'CREDITO' if tipo_conta_excel and 'credito' in str(tipo_conta_excel).lower() else 'CORRENTE'
                
                # Cartao (Excel) → nome_conta real na aplicação
                nome_conta = str(cartao_nome_excel).strip() if cartao_nome_excel else 'Padrão'
                
                # Cria conta se não existe (baseado no tipo)
                if nome_conta not in contas_criadas:
                    nova_conta_id = max([c.get('id', 0) for c in dados['contas']], default=0) + 1
                    dados['contas'].append({
                        'id': nova_conta_id,
                        'nome': nome_conta,
                        'tipo': tipo_conta_normalizado,
                        'saldo_manual': 0,
                        'criada_em': datetime.now().isoformat()
                    })
                    contas_criadas.add(nome_conta)
                
                # Se for cartão de crédito, criar cartão também
                cartao_id = None
                if tipo_conta_normalizado in ['CREDITO', 'CARTAO'] and nome_conta not in cartoes_criados:
                    novo_cartao_id = max([c.get('id', 0) for c in dados['cartoes']], default=0) + 1
                    dados['cartoes'].append({
                        'id': novo_cartao_id,
                        'nome': nome_conta,
                        'bandeira': nome_conta.split()[0] if nome_conta else 'Cartão',
                        'limite': 5000,
                        'criada_em': datetime.now().isoformat()
                    })
                    cartoes_criados.add(nome_conta)
                
                # Processa categorias
                if categoria and str(categoria).strip():
                    categoria_nome = str(categoria).strip()
                    existe_cat = any(c.get('nome') == categoria_nome for c in dados['categorias'])
                    if not existe_cat:
                        nova_cat_id = max([c.get('id', 0) for c in dados['categorias']], default=0) + 1
                        dados['categorias'].append({
                            'id': nova_cat_id,
                            'nome': categoria_nome,
                            'tipo': 'DESPESA' if 'débito' in str(tipo_transacao).lower() else 'RECEITA',
                            'subcategorias': [{'id': 1, 'nome': str(subcategoria).strip()}] if subcategoria else []
                        })
                
                # Normaliza status: PAGO/RECEBIDO = já efetivado
                status_normalizado = str(status).upper().strip() if status else 'EFETIVADO'
                if 'PAGO' in status_normalizado or 'RECEBIDO' in status_normalizado:
                    status_normalizado = 'EFETIVADO'
                
                # Cria transação com novo mapeamento
                transacao = {
                    'id': str(uuid.uuid4()),
                    'data_vencimento': data_vcto_str,
                    'data_transacao': data_str,
                    'mes_pagamento': mes_pgto,
                    'tipo_conta': tipo_conta_normalizado,  # CORRENTE ou CREDITO
                    'nome_conta': nome_conta,  # Nome real da conta/cartão
                    'cartao_id': cartao_id,
                    'tipo': str(tipo_transacao).upper() if tipo_transacao else 'DEBITO',
                    'valor': valor,
                    'parcela': str(parcela_info).upper() if parcela_info else 'À VISTA',
                    'status': status_normalizado,
                    'categoria': str(categoria).strip() if categoria else 'Outras',
                    'subcategoria': str(subcategoria).strip() if subcategoria else 'Geral',
                    'tipo_custo': str(tipo_custo).upper() if tipo_custo and tipo_custo in ['VARIÁVEL', 'FIXO'] else 'VARIÁVEL',
                    'observacoes': str(observacacao) if observacacao else '',
                    'descricao': str(descricao).strip() if descricao else 'Transação importada'
                }
                
                dados['transacoes'].append(transacao)
                transacoes_importadas += 1
                
            except Exception as e:
                print(f"Erro ao processar linha {row_idx}: {str(e)}")
                continue
        
        # Salva dados
        salvar_dados(dados)
        
        return jsonify({
            'mensagem': f'✅ {transacoes_importadas} transações importadas com sucesso!',
            'transacoes_importadas': transacoes_importadas,
            'contas_criadas': len(contas_criadas),
            'cartoes_criados': len(cartoes_criados)
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro ao importar: {str(e)}'}), 400

# ===== ENDPOINTS DE RASTREAMENTO DE SALDOS =====

@app.route('/api/registrar-saldo-diario', methods=['POST'])
def registrar_saldo_diario():
    """Registra o saldo de uma conta em um dia específico"""
    try:
        dados = carregar_dados()
        conta_id = request.json.get('conta_id')
        saldo = float(request.json.get('saldo', 0))
        data = request.json.get('data', datetime.now().strftime('%Y-%m-%d'))
        
        # Encontra a conta
        conta = next((c for c in dados['contas'] if c['id'] == conta_id), None)
        if not conta:
            return jsonify({'erro': 'Conta não encontrada'}), 404
        
        # Inicializa histórico se não existe
        if 'historico_saldos' not in conta:
            conta['historico_saldos'] = []
        
        # Remove saldo anterior da mesma data se existir
        conta['historico_saldos'] = [h for h in conta['historico_saldos'] if h['data'] != data]
        
        # Adiciona novo saldo
        conta['historico_saldos'].append({
            'data': data,
            'saldo': saldo,
            'registrado_em': datetime.now().isoformat()
        })
        
        # Ordena por data
        conta['historico_saldos'].sort(key=lambda x: x['data'])
        
        # Se for hoje, também atualiza saldo_manual para sincronizar com a página principal
        hoje = datetime.now().strftime('%Y-%m-%d')
        if data == hoje:
            conta['saldo_manual'] = saldo
        
        salvar_dados(dados)
        
        return jsonify({
            'mensagem': f'✅ Saldo de R$ {saldo:.2f} registrado para {conta["bandeira"]} em {data}',
            'conta': conta
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/historico-saldos/<int:conta_id>', methods=['GET'])
def obter_historico_saldos(conta_id):
    """Obtém histórico de saldos de uma conta com filtro de período"""
    try:
        dados = carregar_dados()
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        # Encontra a conta
        conta = next((c for c in dados['contas'] if c['id'] == conta_id), None)
        if not conta:
            return jsonify({'erro': 'Conta não encontrada'}), 404
        
        historico = conta.get('historico_saldos', [])
        
        # Filtra por período se fornecido
        if data_inicio:
            historico = [h for h in historico if h['data'] >= data_inicio]
        if data_fim:
            historico = [h for h in historico if h['data'] <= data_fim]
        
        return jsonify({
            'conta': {
                'id': conta['id'],
                'bandeira': conta['bandeira'],
                'tipo': conta['tipo'],
                'saldo_manual': conta.get('saldo_manual', 0)
            },
            'historico': historico,
            'total_registros': len(historico)
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/saldo-consolidado', methods=['GET'])
def obter_saldo_consolidado():
    """Obtém saldo consolidado de todas as contas em uma data específica ou última registrada"""
    try:
        dados = carregar_dados()
        data = request.args.get('data')
        contas_ids = request.args.getlist('contas_ids')  # Filtro opcional de contas
        
        saldos_por_conta = {}
        saldo_total = 0
        
        for conta in dados['contas']:
            # Filtra por IDs se fornecido
            if contas_ids and str(conta['id']) not in contas_ids:
                continue
            
            historico = conta.get('historico_saldos', [])
            
            if data:
                # Busca saldo na data específica
                saldo_dia = next((h['saldo'] for h in historico if h['data'] == data), None)
                if not saldo_dia and not historico:
                    saldo_dia = conta.get('saldo_manual', 0)
            else:
                # Usa último saldo registrado ou saldo_manual
                saldo_dia = historico[-1]['saldo'] if historico else conta.get('saldo_manual', 0)
            
            if saldo_dia is not None:
                saldos_por_conta[conta['id']] = {
                    'bandeira': conta['bandeira'],
                    'tipo': conta['tipo'],
                    'saldo': saldo_dia
                }
                saldo_total += saldo_dia
        
        return jsonify({
            'data': data or 'Última registrada',
            'saldos_por_conta': saldos_por_conta,
            'saldo_consolidado': round(saldo_total, 2),
            'quantidade_contas': len(saldos_por_conta)
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

@app.route('/api/saldo-projetado', methods=['GET'])
def obter_saldo_projetado():
    """Calcula saldo projetado baseado em transações futuras"""
    try:
        dados = carregar_dados()
        contas_ids = request.args.getlist('contas_ids')
        data_ate = request.args.get('data_ate')
        
        if not data_ate:
            data_ate = datetime.now().strftime('%Y-%m-%d')
        
        saldos = {}
        
        for conta in dados['contas']:
            if contas_ids and str(conta['id']) not in contas_ids:
                continue
            
            # Pega último saldo conhecido
            historico = conta.get('historico_saldos', [])
            saldo_base = historico[-1]['saldo'] if historico else conta.get('saldo_manual', 0)
            
            # Adiciona transações EFETIVADAS com data_vencimento até data_ate
            saldo_projetado = saldo_base
            
            for transacao in dados['transacoes']:
                if transacao.get('status') != 'EFETIVADO':
                    continue
                
                data_vcto = transacao.get('data_vencimento', '')
                
                # Verifica se a transação afeta essa conta
                if str(transacao.get('nome_conta', '')).lower() in str(conta['bandeira']).lower() or \
                   transacao.get('tipo_conta') == conta['tipo']:
                    
                    if data_vcto <= data_ate:
                        valor = float(transacao.get('valor', 0))
                        if transacao.get('tipo') in ['DEBITO', 'DESPESA']:
                            saldo_projetado -= valor
                        else:
                            saldo_projetado += valor
            
            saldos[conta['id']] = {
                'bandeira': conta['bandeira'],
                'tipo': conta['tipo'],
                'saldo_base': saldo_base,
                'saldo_projetado': round(saldo_projetado, 2),
                'diferenca': round(saldo_projetado - saldo_base, 2)
            }
        
        saldo_consolidado_projetado = sum(s['saldo_projetado'] for s in saldos.values())
        
        return jsonify({
            'data_projecao': data_ate,
            'saldos_por_conta': saldos,
            'saldo_consolidado_projetado': round(saldo_consolidado_projetado, 2),
            'quantidade_contas': len(saldos)
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

if __name__ == '__main__':
    print("=" * 50)
    print("🚀 SISTEMA DE FINANÇAS INICIADO")
    print("=" * 50)
    print("")
    print("📱 Acesse em seu NAVEGADOR:")
    print("   http://localhost:3333")
    print("")
    print("=" * 50)
    app.run(debug=False, host='127.0.0.1', port=3333, use_reloader=False)
